import requests
import base64
import datetime
import traceback
from openpyxl import load_workbook, Workbook

######################################

# Yes for True otherwise False
isAllCounty = False

Harnett_County = True
Johnston_County = False 
Lee_County = False 
Mecklenburg_County = False 
Wake_County = False 

# Case Type Select (Enter the catagory value or Keep empty for select all)
case_type = 'IF'

# Party Connection Types
partyExtendedConnectionTypes = 'DEF'

# Input Date as MM/DD/YYYY Format
caseFiledStartDate = '10/01/2023'
caseFiledEndDate = '10/05/2023'

######################################

# Output Param
countyNodeIDParam = '?countyNodeID='

# Each county Category Values
Harnett_County_value = ['101043000', '101043001', '101043002', '101043003']
Johnston_County_value = ['101051000', '101051001', '101051002', '101051003']
Lee_County_value = ['101053000', '101053001', '101053002', '101053003']
Mecklenburg_County_value = ['101060000', '101060001', '101060002', '101060003']
Wake_County_value = ['101092000', '101092001', '101092002', '101092003']

# Value addedd if True
if Harnett_County: countyNodeIDParam += '&countyNodeID='.join(Harnett_County_value)
if Johnston_County: countyNodeIDParam += '&countyNodeID='.join(Johnston_County_value)
if Lee_County: countyNodeIDParam += '&countyNodeID='.join(Lee_County_value)
if Mecklenburg_County: countyNodeIDParam += '&countyNodeID='.join(Mecklenburg_County_value)
if Wake_County: countyNodeIDParam += '&countyNodeID='.join(Wake_County_value)

# All Country
if isAllCounty:
    countyNodeIDParam = '?countyNodeID='
    countyNodeIDParam += '&countyNodeID='.join(Harnett_County_value + Johnston_County_value + Lee_County_value + Mecklenburg_County_value + Wake_County_value)

# Case Type added
if len(case_type) != 0:
    countyNodeIDParam += f"&caseType={case_type}"

# Party Connection Types
if len(partyExtendedConnectionTypes) != 0:
    countyNodeIDParam += f"&partyExtendedConnectionTypes={partyExtendedConnectionTypes}"

#Case File Start Data
if len(caseFiledStartDate) != 0:
    countyNodeIDParam += f"&caseFiledStartDate={caseFiledStartDate}"

#Case File end Data
if len(caseFiledEndDate) != 0:
    countyNodeIDParam += f"&caseFiledEndDate={caseFiledEndDate}"

# Write Headline and create a new excel sheet
def xl_sheet_headlines(sheet_name):
    wb = Workbook()
    ws = wb.active
    headlines = headlines = ["caseNumber", "assignmentDate", "caseStatusDate", "caseStatusTypeCode", "caseStatusTypeCodedescription", "caseTitle", "caseType", "courtName", "active", "timestampCreate", "OffenseStatuteDescription", "citationNumber", "offenseStatuteNumber", "citationDegree", "chargeOffenseDate", "chargeOffenseTime", "defendant_Party_firstName", "defendant_Party_middleName", "defendant_Party_lastName", "defendant_Party_address", "defendant_Party_city", "defendant_Party_state", "defendant_Party_zip", "defendant_Party_gender", "defendant_Party_race", "defendant_Party_heightFeet", "defendant_Party_heightInches", "defendant_Party_dateOfBirth", "defendant_Party_ethnicity", "defendant_Party_hairColor", "defendant_Party_eyeColor", "defendant_Party_needsInterpreter", "defendant_Party_partyInJailFlag", "defendant_Party_internalPartyID", "defendant_Party_registeredSexOffenderFlag", "State_Party_Name", "Complainant_Party_firstName", "Complainant_Party_lastName", "Complainant_Party_address", "Complainant_Party_address2", "Complainant_Party_city", "Complainant_Party_state", "Complainant_Party_zip", "Complainant_Party_needsInterpreter", "Complainant_Party_maritalStatus", "Complainant_Party_partyInJailFlag", "Complainant_Party_internalPartyID", "Complainant_Party_registeredSexOffenderFlag", "hearing_hearingDate", "hearing_hearingTime", "hearing_hearingType_code", "hearing_hearingType_description", "hearing_cancelled", "hearing_timestampCreate", "hearing_courtSessionName", "hearing_Hearing_Location_code", "hearing_Hearing_Location_desc", "citee_nameFirst", "citee_nameLast", "citee_description", "citee_offenseDate", "citee_vehicle_licensePlate", "citee_vehicle_state", "citee_vehicle_year", "citee_vehicleMake", "citee_vehicleType", "citee_commercialVehicleFlag", "citee_vehicleInactive", "citee_hazardousVehicleFlag", "citee_fine", "citee_countyLocation", "citee_officerName", "Total_Financial_Assessment", "Total_Payments", "Balance", "Financial_Transaction_Date", "eventType_code", "eventType_description", "caseEventDate", "caseEventCreateDate", "Reporting_Agency_Code", "Reporting_Agency_description", "charges_Office_Name", "charges_Office_BadgeNumber", "case_url"]
    ws.append(headlines)
    wb.save(sheet_name)

# Write Data On existing sheet
def xl_write(data_write, sheet_name):
    wb = load_workbook(sheet_name)
    work_sheet = wb.active # Get active sheet
    work_sheet.append(data_write)
    wb.save(sheet_name)

def get_accessToken(email, password, scope, client_id, auth_url):

    # Encode email and password to base64
    encoded_email = base64.b64encode(email.encode()).decode()
    encoded_password = base64.b64encode(password.encode()).decode()

    # Requesting authentication token
    payload = {
        "appClientId": client_id,
        "tokenScope": scope,
        "userEmailAddress": encoded_email,
        "userPassword": encoded_password
    }

    # Request headers with content type
    headers = {
        "Content-Type": "application/x-www-form-urlencoded"
    }

    response = requests.post(auth_url, data=payload, headers=headers, proxies=proxies)
    data = response.json()

    try:
        accessToken = data['accessToken']
        print('Retrive Access Token sucessfully')
        return accessToken
    except:
        print(traceback.format_exc())
        print('SomeThing Went Wrong, please Try again or Contact with devoloper kawsarlog@gmail.com')
        input('Press Input to continue')

def get_all_cases(accessToken, countyNodeIDParam):

    # cases List
    cases = []

    perPageResultCount = 15

    headers = {
        'authority': 'prdaws.nccourts.org',
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'en-US,en;q=0.9',
        'authorization': f'Bearer {accessToken}',
        'referer': 'https://prdaws.nccourts.org/rpa_ui/cases',
        'sec-ch-ua': '"Chromium";v="118", "Google Chrome";v="118", "Not=A?Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36',
    }

    print('Working at page no> ', end='')
    for page in range(1, 9999999):
        print(page, end='-')
        response = requests.get(
            f'https://prdaws.nccourts.org/rpa_web_services/api/v1/partycases{countyNodeIDParam}&page={page}',
            headers=headers,
            proxies=proxies
        )

        # append all cases
        cases += response.json()['cases']

        Total_results = response.json()['totalResults']

        if (Total_results % perPageResultCount == 0):
            total_page = int(Total_results / perPageResultCount)
        else:
            total_page = int(Total_results / perPageResultCount) + 1

        if page == total_page:
            break

    print(f"Total {len(cases)} cases Found")
    return cases

def getAll_node_case_id(cases):
    
    node_case_ids = []
    case_ids = []
    for case in cases:
        node_id = str(case['nodeID'])
        case_id = str(case['caseID'])
        
        node_case_id = [node_id, case_id]
        if case_id not in case_ids:
            node_case_ids.append(node_case_id)
            case_ids.append(case_id)
    print(f"Unique {len(node_case_ids)} cases Found")
    return node_case_ids

def get_case_response_json(case_url):
    
    global accessToken
    
    headers = {
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'en-US,en;q=0.9',
        'authorization': f'Bearer {accessToken}',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36',
    }

    response = requests.get(case_api_url, headers=headers, proxies=proxies)
    json_response = response.json()

    # If Token Expire Try Again
    if 'message' in json_response:
        is_token_expire = json_response['message'] == 'Access token has expired.'

        if is_token_expire:
            print('Token Expire')
            accessToken = get_accessToken(email, password, scope, client_id, auth_url)
            response = requests.get(case_api_url, headers=headers, proxies=proxies)
            json_response = response.json()

    return json_response

def get_connection_type_data(json_response, partyType):
    caseParties = json_response['caseDetails']['caseParties']
    for casePartie in caseParties:
        connections = casePartie['connection']
        for connection in connections:
            if connection['description']==partyType:
                return casePartie
    return None

def get_connection_type_data2(json_response, partyType):
    for parties in json_response['parties']:
        if partyType in parties:
            return parties
    return None

def get_Complainant_json_ele3(Complainant_json_ele2):
    for i in Complainant_json_ele2['address']:
        if i['partyCurrent']:
            return i
    return {}

proxies = {
    'http': 'http://xxxxxx:xxxxxxxxxx@104.232.209.54:6012',
    'https': 'http://xxxxxx:xxxxxxxxx@104.232.209.54:6012'
}

email = ''
password = ""
client_id = ""

scope = "api://prd-rpa-web-services.nclea.gov/User.ReadAccess"
auth_url = "https://prdaws.nccourts.org/authentication_proxy/api/v1/authorize"

accessToken = get_accessToken(email, password, scope, client_id, auth_url)

try:
    cases = get_all_cases(accessToken, countyNodeIDParam)
except:
    print(traceback.format_exc())
    print('SomeThing Went Wrong, please Try again or Contact with devoloper kawsarlog@gmail.com')
    input('Press Input to continue')

node_case_ids = getAll_node_case_id(cases)

print('Creating New Excel Sheet')

# create filename for this run. Unique date and time for each file.
scrapeDateTime = datetime.datetime.now().strftime("%B_%d_%Y_%H%M")
#csv_filename = 'Vrbo_Data_Manasota_Key_FL_' + scrapeDateTime + '.csv'
sheet_name = 'nccourts_Data_' + scrapeDateTime + '.xlsx'
# Write Headline
xl_sheet_headlines(sheet_name)

for node_case_id in node_case_ids:
    node_id, case_id = node_case_id

    case_url = f"https://prdaws.nccourts.org/rpa_ui/case-detail/{case_id}/{node_id}"
    case_api_url = f'https://prdaws.nccourts.org/rpa_web_services/api/v1/cases/details/{node_id}/{case_id}'
    print(f"Working at {case_url}")

    # Response
    json_response = get_case_response_json(case_url)

    try: Defendant_json_ele = get_connection_type_data(json_response, 'Defendant')
    except: Defendant_json_ele = {}

    try: Defendant_json_ele2 = get_connection_type_data2(json_response, 'gender')
    except: Defendant_json_ele2 = {}

    try: Complainant_json_ele = get_connection_type_data(json_response, 'Complainant')
    except: Complainant_json_ele = {}

    try: Complainant_json_ele2 = get_connection_type_data2(json_response, 'maritalStatus')
    except: Complainant_json_ele2 = {}

    try: Complainant_json_ele3 = get_Complainant_json_ele3(Complainant_json_ele2)
    except: Complainant_json_ele3 = {}

    caseNumber = '' 
    assignmentDate = '' 
    caseStatusDate = '' 
    caseStatusTypeCode = '' 
    caseStatusTypeCodedescription = '' 
    caseTitle = '' 
    caseType = '' 
    courtName = '' 
    active = '' 
    timestampCreate = '' 
    OffenseStatuteDescription = '' 
    citationNumber = '' 
    offenseStatuteNumber = '' 
    citationDegree = '' 
    chargeOffenseDate = '' 
    chargeOffenseTime = '' 
    defendant_Party_firstName = '' 
    defendant_Party_middleName = '' 
    defendant_Party_lastName = '' 
    defendant_Party_address = '' 
    defendant_Party_city = '' 
    defendant_Party_state = '' 
    defendant_Party_zip = '' 
    defendant_Party_gender = '' 
    defendant_Party_race = '' 
    defendant_Party_heightFeet = '' 
    defendant_Party_heightInches = '' 
    defendant_Party_dateOfBirth = '' 
    defendant_Party_ethnicity = '' 
    defendant_Party_hairColor = '' 
    defendant_Party_eyeColor = '' 
    defendant_Party_needsInterpreter = '' 
    defendant_Party_partyInJailFlag = '' 
    defendant_Party_internalPartyID = '' 
    defendant_Party_registeredSexOffenderFlag = '' 
    State_Party_Name = '' 
    Complainant_Party_firstName = '' 
    Complainant_Party_lastName = '' 
    Complainant_Party_address = '' 
    Complainant_Party_address2 = '' 
    Complainant_Party_city = '' 
    Complainant_Party_state = '' 
    Complainant_Party_zip = '' 
    Complainant_Party_needsInterpreter = '' 
    Complainant_Party_maritalStatus = '' 
    Complainant_Party_partyInJailFlag = '' 
    Complainant_Party_internalPartyID = '' 
    Complainant_Party_registeredSexOffenderFlag = '' 
    hearing_hearingDate = '' 
    hearing_hearingTime = '' 
    hearing_hearingType_code = '' 
    hearing_hearingType_description = '' 
    hearing_cancelled = '' 
    hearing_timestampCreate = '' 
    hearing_courtSessionName = '' 
    hearing_Hearing_Location_code = '' 
    hearing_Hearing_Location_desc = '' 
    citee_nameFirst = '' 
    citee_nameLast = '' 
    citee_description = '' 
    citee_offenseDate = '' 
    citee_vehicle_licensePlate = '' 
    citee_vehicle_state = '' 
    citee_vehicle_year = '' 
    citee_vehicleMake = '' 
    citee_vehicleType = '' 
    citee_commercialVehicleFlag = '' 
    citee_vehicleInactive = '' 
    citee_hazardousVehicleFlag = '' 
    citee_fine = '' 
    citee_countyLocation = '' 
    citee_officerName = '' 
    Total_Financial_Assessment = '' 
    Total_Payments = '' 
    Balance = '' 
    Financial_Transaction_Date = '' 
    eventType_code = '' 
    eventType_description = '' 
    caseEventDate = '' 
    caseEventCreateDate = '' 
    Reporting_Agency_Code = '' 
    Reporting_Agency_description = '' 
    charges_Office_Name = '' 
    charges_Office_BadgeNumber = '' 

    try: caseNumber = json_response['caseDetails']['caseNumber']
    except: pass
    print('caseNumber:', caseNumber)
    
    try: assignmentDate = json_response['caseDetails']['assignments'][0]['assignmentDate']
    except: pass
    try: caseStatusDate = json_response['caseDetails']['caseStatus'][0]['caseStatusDate']
    except: pass
    try: caseStatusTypeCode = json_response['caseDetails']['caseStatus'][0]['caseStatusType']['code']
    except: pass
    try: caseStatusTypeCodedescription = json_response['caseDetails']['caseStatus'][0]['caseStatusType']['description']
    except: pass
    try: caseTitle = json_response['caseDetails']['caseTitle']
    except: pass
    try: caseType = json_response['caseDetails']['caseType']['description']
    except: pass
    try: courtName = json_response['caseDetails']['court']['courtName']
    except: pass
    try: active = json_response['caseDetails']['active']
    except: pass
    try: timestampCreate = json_response['caseDetails']['caseStatus'][0]['timestampCreate']
    except: pass
    try: OffenseStatuteDescription = json_response['caseDetails']['charges'][0]['chargeHistories'][0]['statute']['statuteDescription']
    except: pass
    try: citationNumber = json_response['citation']['citationNumber']
    except: pass
    try: offenseStatuteNumber = json_response['caseDetails']['charges'][0]['chargeHistories'][0]['statute']['statuteNumber']
    except: pass
    try: citationDegree = json_response['caseDetails']['charges'][0]['chargeHistories'][0]['statute']['degree']['code']
    except: pass
    try: chargeOffenseDate = json_response['caseDetails']['charges'][0]['chargeOffenseDate']
    except: pass
    try: chargeOffenseTime = json_response['caseDetails']['charges'][0]['chargeOffenseTime']
    except: pass

    # defendant_Party
    try: defendant_Party_firstName = Defendant_json_ele['casePartyName'][0]['firstName']
    except: pass
    try: defendant_Party_middleName = Defendant_json_ele['casePartyName'][0]['middleName']
    except: pass
    try: defendant_Party_lastName = Defendant_json_ele['casePartyName'][0]['lastName']
    except: pass
    try: defendant_Party_address = Defendant_json_ele['casePartyAddress']['addressLine1']
    except: pass
    try: defendant_Party_city = Defendant_json_ele['casePartyAddress']['city']
    except: pass
    try: defendant_Party_state = Defendant_json_ele['casePartyAddress']['state']
    except: pass
    try: defendant_Party_zip = Defendant_json_ele['casePartyAddress']['zip']
    except: pass
    try: defendant_Party_gender = Defendant_json_ele2['gender']
    except: pass
    try: defendant_Party_race = Defendant_json_ele2['race']
    except: pass
    try: defendant_Party_heightFeet = Defendant_json_ele2['heightFeet']
    except: pass
    try: defendant_Party_heightInches = Defendant_json_ele2['heightInches']
    except: pass
    try: defendant_Party_dateOfBirth = Defendant_json_ele['dateOfBirth']
    except: pass
    try: defendant_Party_ethnicity = Defendant_json_ele2['ethnicity']
    except: pass
    try: defendant_Party_hairColor = Defendant_json_ele2['hairColor']
    except: pass
    try: defendant_Party_eyeColor = Defendant_json_ele2['eyeColor']
    except: pass
    try: defendant_Party_needsInterpreter = Defendant_json_ele2['needsInterpreter']
    except: pass
    try: defendant_Party_partyInJailFlag = Defendant_json_ele2['partyInJailFlag']
    except: pass
    try: defendant_Party_internalPartyID = Defendant_json_ele2['internalPartyID']
    except: pass
    try: defendant_Party_registeredSexOffenderFlag = Defendant_json_ele2['registeredSexOffenderFlag']
    except: pass

    # State_Party
    try: State_Party_Name = get_connection_type_data(json_response, 'State')['casePartyName'][0]['formattedName']
    except: pass

    # Complainant_Party
    try: Complainant_Party_firstName = Complainant_json_ele['casePartyName'][0]['firstName']
    except: pass
    try: Complainant_Party_lastName = Complainant_json_ele['casePartyName'][0]['lastName']
    except: pass
    try: Complainant_Party_address = Complainant_json_ele3['addressLine1']
    except: pass
    try: Complainant_Party_address2 = Complainant_json_ele3['addressLine2']
    except: pass
    try: Complainant_Party_city = Complainant_json_ele3['city']
    except: pass
    try: Complainant_Party_state = Complainant_json_ele3['state']
    except: pass
    try: Complainant_Party_zip = Complainant_json_ele3['zip']
    except: pass
    try: Complainant_Party_needsInterpreter = Complainant_json_ele2['needsInterpreter']
    except: pass
    try: Complainant_Party_maritalStatus = Complainant_json_ele2['maritalStatus']['description']
    except: pass
    try: Complainant_Party_partyInJailFlag = Complainant_json_ele2['partyInJailFlag']
    except: pass
    try: Complainant_Party_internalPartyID = Complainant_json_ele2['internalPartyID']
    except: pass
    try: Complainant_Party_registeredSexOffenderFlag = Complainant_json_ele2['registeredSexOffenderFlag']
    except: pass

    # hearings
    try: hearing_hearingDate = json_response['caseDetails']['hearings'][0]['setting']['hearingDate']
    except: pass
    try: hearing_hearingTime = json_response['caseDetails']['hearings'][0]['setting']['courtSessionBlock']['startTime']
    except: pass
    try: hearing_hearingType_code = json_response['caseDetails']['hearings'][0]['hearingType']['code']
    except: pass
    try: hearing_hearingType_description = json_response['caseDetails']['hearings'][0]['hearingType']['description']
    except: pass
    try: hearing_cancelled = json_response['caseDetails']['hearings'][0]['setting']['cancelled']
    except: pass
    try: hearing_timestampCreate = json_response['caseDetails']['hearings'][0]['timestampCreate']
    except: pass
    try: hearing_courtSessionName = json_response['caseDetails']['hearings'][0]['setting']['courtSessionName']
    except: pass
    try: hearing_Hearing_Location_code = json_response['caseDetails']['hearings'][0]['setting']['courtResource'][0]['code']['code']
    except: pass
    try: hearing_Hearing_Location_desc = json_response['caseDetails']['hearings'][0]['setting']['courtResource'][0]['code']['description']
    except: pass

    # citation
    try: citee_nameFirst = json_response['citation']['citee']['citeeName']['nameFirst']
    except: pass
    try: citee_nameLast = json_response['citation']['citee']['citeeName']['nameLast']
    except: pass
    try: citee_description = json_response['citation']['caseTypeKey']['description']
    except: pass
    try: citee_offenseDate = json_response['citation']['offenseDate']
    except: pass
    try: citee_vehicle_licensePlate = json_response['citation']['vehicle']['licensePlate']
    except: pass
    try: citee_vehicle_state = json_response['citation']['vehicle']['licenseState']['code']
    except: pass
    try: citee_vehicle_year = json_response['citation']['vehicle']['year']
    except: pass
    try: citee_vehicleMake = json_response['citation']['vehicle']['vehicleMake']['description']
    except: pass
    try: citee_vehicleType = json_response['citation']['vehicle']['vehicleType']['description']
    except: pass
    try: citee_commercialVehicleFlag = json_response['citation']['vehicle']['commercialVehicleFlag']
    except: pass
    try: citee_vehicleInactive = json_response['citation']['vehicle']['vehicleInactive']
    except: pass
    try: citee_hazardousVehicleFlag = json_response['citation']['vehicle']['hazardousVehicleFlag']
    except: pass
    try: citee_fine = json_response['citation']['citationCharges'][0]['fine']
    except: pass
    try: citee_countyLocation = json_response['citation']['incident']['countyLocation']['description']
    except: pass
    try: citee_officerName = json_response['citation']['incident']['officerName']
    except: pass

    # Financial
    try: Total_Financial_Assessment = json_response['feesDetails']['feeCategoryBalances']['assessmentBalance']['balance']
    except: pass
    try: Total_Payments = sum([float(line['paymentAmount']) for line in json_response['feesDetails']['feeInstances']['feeInstance']])
    except: pass
    try: Balance = float(Total_Financial_Assessment) - Total_Payments
    except: pass
    try: Financial_Transaction_Date = json_response['feesDetails']['transactions']['transaction'][0]['transactionDate']
    except: pass

    # Event
    try: eventType_code = json_response['caseDetails']['caseEvents'][0]['eventType']['code']
    except: pass
    try: eventType_description =  json_response['caseDetails']['caseEvents'][0]['eventType']['description']
    except: pass
    try: caseEventDate = json_response['caseDetails']['caseEvents'][0]['caseEventDate']
    except: pass
    try: caseEventCreateDate = json_response['caseDetails']['caseEvents'][0]['timestampCreate']
    except: pass

    # Charges
    try: Reporting_Agency_Code = json_response['caseDetails']['charges'][0]['reportingAgency']['agency']['code']
    except: pass
    try: Reporting_Agency_description = json_response['caseDetails']['charges'][0]['reportingAgency']['agency']['description']
    except: pass
    try: charges_Office_Name = json_response['caseDetails']['charges'][0]['reportingAgency']['officerName']
    except: pass
    try: charges_Office_BadgeNumber = json_response['caseDetails']['charges'][0]['reportingAgency']['officerBadgeNumber']
    except: pass

    write_data = [caseNumber, assignmentDate, caseStatusDate, caseStatusTypeCode, caseStatusTypeCodedescription, caseTitle, caseType, courtName, active, timestampCreate, OffenseStatuteDescription, citationNumber, offenseStatuteNumber, citationDegree, chargeOffenseDate, chargeOffenseTime, defendant_Party_firstName, defendant_Party_middleName, defendant_Party_lastName, defendant_Party_address, defendant_Party_city, defendant_Party_state, defendant_Party_zip, defendant_Party_gender, defendant_Party_race, defendant_Party_heightFeet, defendant_Party_heightInches, defendant_Party_dateOfBirth, defendant_Party_ethnicity, defendant_Party_hairColor, defendant_Party_eyeColor, defendant_Party_needsInterpreter, defendant_Party_partyInJailFlag, defendant_Party_internalPartyID, defendant_Party_registeredSexOffenderFlag, State_Party_Name, Complainant_Party_firstName, Complainant_Party_lastName, Complainant_Party_address, Complainant_Party_address2, Complainant_Party_city, Complainant_Party_state, Complainant_Party_zip, Complainant_Party_needsInterpreter, Complainant_Party_maritalStatus, Complainant_Party_partyInJailFlag, Complainant_Party_internalPartyID, Complainant_Party_registeredSexOffenderFlag, hearing_hearingDate, hearing_hearingTime, hearing_hearingType_code, hearing_hearingType_description, hearing_cancelled, hearing_timestampCreate, hearing_courtSessionName, hearing_Hearing_Location_code, hearing_Hearing_Location_desc, citee_nameFirst, citee_nameLast, citee_description, citee_offenseDate, citee_vehicle_licensePlate, citee_vehicle_state, citee_vehicle_year, citee_vehicleMake, citee_vehicleType, citee_commercialVehicleFlag, citee_vehicleInactive, citee_hazardousVehicleFlag, citee_fine, citee_countyLocation, citee_officerName, Total_Financial_Assessment, Total_Payments, Balance, Financial_Transaction_Date, eventType_code, eventType_description, caseEventDate, caseEventCreateDate, Reporting_Agency_Code, Reporting_Agency_description, charges_Office_Name, charges_Office_BadgeNumber, case_url]

    # Write on Excel Sheet
    xl_write(write_data, sheet_name)
